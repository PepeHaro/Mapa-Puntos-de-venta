#!/usr/bin/env python3
"""Da de alta en la BD los distribuidores que solo trae Urrea.

Uso:
    python3 altas_urrea.py            # muestra que agregaria
    python3 altas_urrea.py --aplicar  # escribe en Distribuidores.xlsx

Urrea lista 1,823 distribuidores, casi seis veces mas que los 363 de la BD,
pero 1,552 de ellos tienen una sola sucursal: son ferreterias y plomerias de
barrio, otro segmento que los distribuidores de acabados del padron. Por eso
solo entran los que tienen dos o mas sucursales.

Cuatro filtros, en orden:

1. El nombre ya esta en la BD.
2. Una sola sucursal.
3. El grueso de sus sucursales cae en direcciones que la BD ya tiene a nombre
   de otra empresa. Urrea escribe la razon social de cada socio y la BD el
   nombre del grupo: 142 tiendas que la BD llama GERSA aparecen aqui con el
   nombre legal de cada miembro.
4. Sucursal por sucursal, si comparte calle y numero exterior con una tienda
   de la BD, esa no se da de alta aunque el resto de la cadena si.
"""

import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SERVIDOR = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales"
)
BD = SERVIDOR / "Distribuidores.xlsx"
URREA = SERVIDOR / "Marcas" / "Urrea.xlsx"

MIN_SUCURSALES = 2   # con una sola no entra
CERCA_M = 150        # solo se comparan direcciones dentro de este radio
# Si esta fraccion de sus tiendas cae en direcciones que la BD tiene a nombre
# de otra empresa, es esa empresa. Se usa 0.30 y no la mitad porque la marca
# solo conoce una parte de la red: de las 208 tiendas que Urrea llama
# "MUEBLES PARA BAÑO", 89 estan en direcciones que la BD registra como GERSA,
# y con la mitad como corte se habrian dado de alta como cadena aparte.
PARTE_DEL_GRUPO = 0.30

RUIDO = {"grupo", "comercializadora", "distribuidora", "distribuidor",
         "corporativo", "sa", "cv", "s", "rl", "sapi", "gpo", "de"}

VIAS = {"av", "avenida", "blvd", "boulevard", "calle", "calz", "calzada", "carr",
        "carretera", "prol", "prolongacion", "col", "colonia", "int", "num",
        "numero", "esq", "local", "lote", "sur", "norte", "ote", "pte",
        "centro", "fracc", "esquina"}


def na(s):
    s = "".join(c for c in unicodedata.normalize("NFD", str(s or "").lower())
                if unicodedata.category(c) != "Mn")
    return " ".join(re.sub(r"[^\w\s]", " ", s).split())


def clave(s):
    return " ".join(sorted(w for w in na(s).split() if w not in RUIDO and len(w) > 2))


def sin_cp(direccion):
    """Quita el codigo postal, que si no se cuela como numero de la calle."""
    t = na(direccion)
    t = re.sub(r"\bc\s*p\s*:?\s*\d{4,5}\b", " ", t)
    return " ".join(re.sub(r"\b\d{5}\b", " ", t).split())


def numero_exterior(direccion):
    nums = [int(n) for n in re.findall(r"\b(\d{1,5})\b", sin_cp(direccion))]
    return max(nums) if nums else None


def calle(direccion):
    return {w for w in re.sub(r"\d+", " ", sin_cp(direccion)).split()
            if len(w) > 3 and w not in VIAS}


def misma_tienda(dir_a, dir_b):
    na_, nb = numero_exterior(dir_a), numero_exterior(dir_b)
    if na_ is None or nb is None or na_ != nb:
        return False
    return bool(calle(dir_a) & calle(dir_b))


def km(a, b, c, d):
    return (((a - c) * 110.57) ** 2
            + ((b - d) * 111.32 * math.cos(math.radians(a))) ** 2) ** 0.5


def main():
    aplicar = "--aplicar" in sys.argv

    libro = openpyxl.load_workbook(BD)  # sin data_only: conserva las formulas
    hoja = libro["BD"]
    existentes, puntos = set(), defaultdict(list)
    for fila in hoja.iter_rows(min_row=2):
        if not fila[0].value:
            continue
        existentes.add(clave(fila[0].value))
        try:
            lat, lon = float(fila[7].value), float(fila[8].value)
        except (TypeError, ValueError):
            continue
        puntos[(round(lat, 2), round(lon, 2))].append(
            (lat, lon, str(fila[6].value or ""), str(fila[0].value))
        )

    wb = openpyxl.load_workbook(URREA, read_only=True, data_only=True)
    urrea = [list(f) for f in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True)]
    wb.close()

    por_dist = defaultdict(list)
    for f in urrea:
        por_dist[f[0]].append(f)

    def vecinas(f):
        lat, lon = f[6], f[7]
        return [(d, n) for dla in (-0.01, 0, 0.01) for dlo in (-0.01, 0, 0.01)
                for x, y, d, n in puntos.get((round(lat + dla, 2), round(lon + dlo, 2)), [])
                if km(lat, lon, x, y) * 1000 < CERCA_M]

    ya_nombre = [d for d in por_dist if clave(d) in existentes]
    resto = {d: v for d, v in por_dist.items() if clave(d) not in existentes}
    chicos = {d: v for d, v in resto.items() if len(v) < MIN_SUCURSALES}
    cadenas = {d: v for d, v in resto.items() if len(v) >= MIN_SUCURSALES}

    # Filtro 3: la cadena en realidad es una empresa que la BD ya tiene
    del_grupo, candidatos = {}, {}
    for d, v in cadenas.items():
        duenos = Counter()
        for f in v:
            for direccion, nombre in vecinas(f):
                if misma_tienda(f[5], direccion):
                    duenos[nombre] += 1
                    break
        if duenos and duenos.most_common(1)[0][1] >= max(2, len(v) * PARTE_DEL_GRUPO):
            del_grupo[d] = (duenos.most_common(1)[0][0], duenos.most_common(1)[0][1], len(v))
        else:
            candidatos[d] = v

    # Filtro 4: sucursales sueltas que la BD ya tiene
    altas, repetidas = [], 0
    for d, v in candidatos.items():
        for f in v:
            if any(misma_tienda(f[5], direccion) for direccion, _ in vecinas(f)):
                repetidas += 1
            else:
                altas.append(f)

    print(f"  distribuidores de Urrea:               {len(por_dist):>5}")
    print(f"  1. el nombre ya esta en la BD:         {len(ya_nombre):>5}")
    print(f"  2. una sola sucursal, no entran:       {len(chicos):>5}")
    print(f"  3. son una empresa que la BD ya tiene: {len(del_grupo):>5}")
    print(f"  4. sucursales sueltas ya registradas:  {repetidas:>5}")
    print(f"\n  ALTAS: {len(altas)} sucursales de {len({f[0] for f in altas})} distribuidores")

    if del_grupo:
        print("\n  los que resultaron ser otra empresa:")
        for d, (nombre, n, tot) in sorted(del_grupo.items(), key=lambda x: -x[1][1])[:15]:
            print(f"     {d[:38]:40} -> {nombre[:24]:26} {n}/{tot}")

    if not aplicar:
        print("\n  Nada se escribio. Para aplicarlo:  python3 altas_urrea.py --aplicar")
        return

    n = hoja.max_row
    for f in altas:
        n += 1
        hoja.cell(n, 1, f[0])   # DISTRIBUIDOR
        hoja.cell(n, 2, f[2])   # NOMBRE
        hoja.cell(n, 4, f[1])   # PAÍS
        hoja.cell(n, 5, f[3])   # ESTADO
        hoja.cell(n, 6, f[4])   # CIUDAD
        hoja.cell(n, 7, f[5])   # DIRECCIÓN
        hoja.cell(n, 8, f[6])   # LATITUD
        hoja.cell(n, 9, f[7])   # LONGITUD
        hoja.cell(n, 10, f'=IF(H{n}="","","https://www.google.com/maps?q="&H{n}&","&I{n})')
    libro.save(BD)
    print(f"\n  BD queda en {hoja.max_row - 1} filas")


if __name__ == "__main__":
    main()
