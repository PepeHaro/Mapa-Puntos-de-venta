#!/usr/bin/env python3
"""Quita de la BD la misma tienda registrada dos veces.

Uso:
    python3 limpiar_bd.py            # muestra que quitaria
    python3 limpiar_bd.py --aplicar  # borra las filas

Dos cosas hacen que una tienda aparezca duplicada:

1. El nombre del distribuidor trae un espacio de mas al final. "PISOS IBERIA" y
   "PISOS IBERIA " se ven identicos pero cuentan como dos empresas, y sus
   sucursales nunca se comparan entre si.
2. La direccion no trae numero. Media Tabasco esta en calles "s/n", asi que
   SANTANDREU tenia 35 filas para sus 18 tiendas: cada una registrada una vez
   con el nombre de la sucursal y otra con el nombre de la cadena.

De cada par se conserva la fila mas completa: la que trae numero de tienda y un
nombre de sucursal propio en vez de repetir el del distribuidor.
"""

import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

BD = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales/Distribuidores.xlsx"
)

# Palabras que aparecen en cualquier direccion y no distinguen una calle de
# otra. Los puntos cardinales importan: sin ellos "11a Norte Oriente" y "17
# Oriente Norte", que son calles distintas de Tapachula, parecian la misma.
VIAS = {"av", "avenida", "blvd", "boulevard", "calle", "calz", "calzada", "carr",
        "carretera", "prol", "prolongacion", "col", "colonia", "int", "num",
        "numero", "esq", "local", "lote", "centro", "fracc", "frac", "esquina",
        "sin", "interior", "piso", "plaza", "sur", "norte", "ote", "pte",
        "oriente", "poniente", "nte", "ste", "esq"}


def na(s):
    s = "".join(c for c in unicodedata.normalize("NFD", str(s or "").lower())
                if unicodedata.category(c) != "Mn")
    return " ".join(re.sub(r"[^\w\s]", " ", s).split())


# Los dos primeros digitos del codigo postal dicen el estado. Es la forma
# segura de distinguirlo del numero de la calle: en Tijuana el Boulevard
# Salinas llega al 11294, que empieza en 11 y seria Ciudad de Mexico.
CP_ESTADO = {}
for _ini, _fin, _edo in [
    (1, 16, "cdmx"), (20, 20, "aguascalientes"), (21, 22, "baja california"),
    (23, 23, "baja california sur"), (24, 24, "campeche"), (25, 27, "coahuila"),
    (28, 28, "colima"), (29, 30, "chiapas"), (31, 33, "chihuahua"),
    (34, 35, "durango"), (36, 38, "guanajuato"), (39, 41, "guerrero"),
    (42, 43, "hidalgo"), (44, 49, "jalisco"), (50, 57, "estado de mexico"),
    (58, 61, "michoacan"), (62, 62, "morelos"), (63, 63, "nayarit"),
    (64, 67, "nuevo leon"), (68, 71, "oaxaca"), (72, 75, "puebla"),
    (76, 76, "queretaro"), (77, 77, "quintana roo"), (78, 79, "san luis potosi"),
    (80, 82, "sinaloa"), (83, 85, "sonora"), (86, 86, "tabasco"),
    (87, 89, "tamaulipas"), (90, 90, "tlaxcala"), (91, 96, "veracruz"),
    (97, 97, "yucatan"), (98, 99, "zacatecas"),
]:
    for _n in range(_ini, _fin + 1):
        CP_ESTADO[_n] = _edo

ESTADOS_MX = set(na(
    "Aguascalientes Baja California Sur Campeche Coahuila Zaragoza Colima Chiapas "
    "Chihuahua Ciudad Mexico CDMX Durango Guanajuato Guerrero Hidalgo Jalisco "
    "Michoacan Ocampo Morelos Nayarit Nuevo Leon Oaxaca Puebla Queretaro Quintana "
    "Roo Luis Potosi Sinaloa Sonora Tabasco Tamaulipas Tlaxcala Veracruz Ignacio "
    "Llave Yucatan Zacatecas"
).split())


def sin_cp(direccion, ciudad="", estado=""):
    """Quita el codigo postal sin llevarse el numero de la calle.

    No sirve borrar todos los numeros de cinco cifras: en Tijuana el Boulevard
    Salinas llega al 11294, y al quitarlo dos locales distintos se quedaban sin
    numero y parecian el mismo. Tampoco sirve quitar el ultimo, porque ese 11294
    es el unico numero de su direccion.

    Lo que si distingue: al codigo postal le sigue el nombre del municipio o del
    estado ("86500 Cárdenas, Tabasco"), y al numero de la calle le sigue la
    colonia ("11264, Fracc. Aviación").
    """
    t = na(direccion)
    t = re.sub(r"\bc\s*p\s*:?\s*\d{4,5}\b", " ", t)
    esperado = na(estado)
    palabras = t.split()
    fuera = []
    for w in palabras:
        if len(w) == 5 and w.isdigit() and CP_ESTADO.get(int(w[:2])) == esperado:
            continue
        fuera.append(w)
    return " ".join(fuera)


def numero_exterior(direccion, ciudad="", estado=""):
    nums = [int(n) for n in re.findall(r"\b(\d{1,5})\b", sin_cp(direccion, ciudad, estado))]
    return max(nums) if nums else None


def calle(direccion, ciudad="", estado=""):
    """Palabras que de verdad identifican la calle.

    Se quitan la ciudad y el estado porque unas direcciones los traen y otras
    no: "Galeana s/n, Cárdenas, Tabasco" y "Galeana S/N, Col. Centro, CP 86500"
    son la misma tienda, y lo unico que comparten es "galeana".
    """
    fuera = VIAS | ESTADOS_MX | set(na(ciudad).split()) | set(na(estado).split())
    return {w for w in re.sub(r"\d+", " ", sin_cp(direccion, ciudad, estado)).split()
            if len(w) > 3 and w not in fuera}


# Cada fuente geocodifico por su cuenta, asi que la misma tienda puede aparecer
# con coordenadas separadas por kilometros: hay filas con la direccion escrita
# igual, letra por letra, a 2.8 km una de otra. Por eso manda la direccion y la
# distancia solo acota.
LEJOS_KM = 3.0


def misma_tienda(a, b, metros):
    """Decide si dos filas son el mismo local. Cada una es (direccion, ciudad, estado).

    Cuando la direccion coincide de verdad, la distancia deja de importar: hay
    filas con la calle y el numero escritos igual a cientos de kilometros una de
    otra porque a una le pusieron mal el municipio. La misma tienda de
    SANTANDREU aparecia en Campeche y en Villahermosa.

    La distancia sigue mandando cuando lo unico compartido es una calle de
    nombre comun, para no juntar dos tiendas distintas de una misma cadena que
    esten en un "Av. Juárez 100" de dos ciudades.
    """
    comun = calle(*a) & calle(*b)
    if not comun:
        return False
    misma_ciudad = na(a[1]) == na(b[1]) and na(a[1]) != ""
    na_, nb = numero_exterior(*a), numero_exterior(*b)
    if na_ is not None and nb is not None:
        if na_ != nb:
            return False
        if len(comun) >= 2 or misma_ciudad:
            return True
        return metros < LEJOS_KM * 1000
    if na_ is None and nb is None:
        if len(comun) >= 3:
            return True
        if len(comun) >= 2:
            return misma_ciudad or metros < LEJOS_KM * 1000
        return metros < 800
    return len(comun) >= 2 and metros < LEJOS_KM * 1000


GENERICAS_SUC = {"suc", "sucursal", "matriz", "bodega", "cedis", "tienda", "pdv",
                 "local", "planta", "centro", "principal"}


def clave_sucursal(nombre, distribuidor):
    """El nombre de la sucursal reducido a lo que la distingue.

    "SUC. ECHEGARAY" y "EL SURTIDOR - ECHEGARAY" son la misma tienda. Se quitan
    las palabras del distribuidor y las de relleno, y se conservan los numeros y
    numerales romanos, que si distinguen: "NIÑOS HÉROES I" no es "NIÑOS HÉROES
    II".
    """
    fuera = set(na(distribuidor).split()) | GENERICAS_SUC | VIAS
    return frozenset(w for w in na(nombre).split() if w and w not in fuera)


def km(a, b, c, d):
    return (((a - c) * 110.57) ** 2
            + ((b - d) * 111.32 * math.cos(math.radians(a))) ** 2) ** 0.5


def riqueza(fila):
    """Que tan completa esta la fila, para saber cual del par conservar."""
    dist = " ".join(str(fila[0].value or "").split()).upper()
    nombre = " ".join(str(fila[1].value or "").split()).upper()
    puntos = sum(1 for c in fila[:9] if c.value not in (None, ""))
    if fila[2].value:                      # trae numero de tienda
        puntos += 2
    if nombre and nombre != dist:          # nombre propio de sucursal
        puntos += 3
    puntos += len(str(fila[6].value or "")) / 200   # direccion mas detallada
    return puntos


def main():
    aplicar = "--aplicar" in sys.argv

    libro = openpyxl.load_workbook(BD)
    hoja = libro["BD"]

    filas, espacios = [], 0
    for fila in hoja.iter_rows(min_row=2):
        if not fila[0].value:
            continue
        limpio = " ".join(str(fila[0].value).split())
        if limpio != str(fila[0].value):
            espacios += 1
        try:
            lat, lon = float(fila[7].value), float(fila[8].value)
        except (TypeError, ValueError):
            continue
        filas.append((fila, limpio, lat, lon, fila[0].row))

    print(f"  BD: {hoja.max_row - 1} filas")
    print(f"  nombres con espacios de sobra: {espacios}")

    por_dist = defaultdict(list)
    for f in filas:
        por_dist[f[1]].append(f)

    fuera, pares = set(), []
    for dist, grupo in por_dist.items():
        for i in range(len(grupo)):
            a = grupo[i]
            if a[4] in fuera:
                continue
            for j in range(i + 1, len(grupo)):
                b = grupo[j]
                if b[4] in fuera:
                    continue
                metros = km(a[2], a[3], b[2], b[3]) * 1000

                def señas(f):
                    return (str(f[0][6].value or ""), str(f[0][5].value or ""),
                            str(f[0][4].value or ""))

                if not misma_tienda(señas(a), señas(b), metros):
                    continue
                queda, sobra = (a, b) if riqueza(a[0]) >= riqueza(b[0]) else (b, a)
                fuera.add(sobra[4])
                pares.append((dist, queda, sobra, metros))

    # Segunda pasada: el nombre de la sucursal. Vale cuando la direccion no
    # alcanza, que pasa cuando una fila no trae numero o trae mal el municipio:
    # una "SUC. TOLUCA" estaba puesta en Tecámac, a 78 km de la de Toluca.
    for dist, grupo in por_dist.items():
        claves = defaultdict(list)
        for f in grupo:
            if f[4] in fuera:
                continue
            k = clave_sucursal(f[0][1].value, dist)
            if k:
                claves[k].append(f)
        for k, iguales in claves.items():
            # Si tres o mas se llaman igual, el nombre no distingue: siete
            # tiendas de DAM en Guadalajara comparten nombre y son distintas.
            if len(iguales) != 2:
                continue
            a_, b_ = iguales
            queda, sobra = (a_, b_) if riqueza(a_[0]) >= riqueza(b_[0]) else (b_, a_)
            fuera.add(sobra[4])
            pares.append((dist, queda, sobra,
                          km(a_[2], a_[3], b_[2], b_[3]) * 1000))

    print(f"  tiendas registradas dos veces: {len(pares)}\n")
    for dist, cuenta in Counter(d for d, _, _, _ in pares).most_common(20):
        print(f"     {dist[:44]:46} {cuenta}")

    print("\n  ejemplos de lo que se quitaria:")
    for dist, queda, sobra, metros in pares[:8]:
        print(f"     {dist[:34]:36} {int(metros)} m")
        print(f"        queda  fila {queda[4]:>5}  {str(queda[0][1].value)[:22]:24} {str(queda[0][6].value)[:52]}")
        print(f"        se va  fila {sobra[4]:>5}  {str(sobra[0][1].value)[:22]:24} {str(sobra[0][6].value)[:52]}")

    if not aplicar:
        print("\n  Nada se escribio. Para aplicarlo:  python3 limpiar_bd.py --aplicar")
        return

    for fila in hoja.iter_rows(min_row=2):
        if fila[0].value:
            fila[0].value = " ".join(str(fila[0].value).split())

    for n in sorted(fuera, reverse=True):
        hoja.delete_rows(n)

    # Las formulas de la columna J traen el numero de fila; hay que rehacerlas.
    for n in range(2, hoja.max_row + 1):
        hoja.cell(n, 10, f'=IF(H{n}="","","https://www.google.com/maps?q="&H{n}&","&I{n})')

    libro.save(BD)
    print(f"\n  se quitaron {len(fuera)} filas; la BD queda en {hoja.max_row - 1}")


if __name__ == "__main__":
    main()
