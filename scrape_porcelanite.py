#!/usr/bin/env python3
"""Genera Marcas/Porcelanite.xlsx desde el localizador de Porcelanite.

Uso:
    python3 scrape_porcelanite.py

La pagina https://porcelanite.com.mx/donde-comprar/ usa el plugin WP Google Map
Gold, que deja los 1,122 puntos de venta incrustados en la variable
`wpgmp.mapdata1`, codificada en base64. Se descarga la pagina una sola vez y se
lee ese bloque.

El sitio responde 403 a curl normal: filtra por huella TLS, no por User-Agent.
Se usa curl_cffi, que imita la de Chrome.

El archivo sale con las mismas columnas que Daltile.xlsx, Interceramic.xlsx,
Cesantoni.xlsx y Vitromex.xlsx, para que generar.py lo lea sin cambios.
"""

import base64
import json
import math
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

URL = "https://porcelanite.com.mx/donde-comprar/"
AQUI = Path(__file__).resolve().parent
DESTINO = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales/Marcas/Porcelanite.xlsx"
)
MARCA_GENERADO = AQUI / ".porcelanite-generado"

COLUMNAS = [
    "DISTRIBUIDOR", "PAÍS", "NOMBRE", "ESTADO", "CIUDAD",
    "DIRECCIÓN", "LATITUD", "LONGITUD", "URL DE UBICACIÓN",
]

# Porcelanite abrevia el estado; la BD lo escribe completo.
ESTADOS = {
    "AGS": "Aguascalientes", "BC": "Baja California", "BCS": "Baja California Sur",
    "CHI": "Chihuahua", "CHS": "Chiapas", "CMP": "Campeche", "CMX": "CDMX",
    "COA": "Coahuila", "COL": "Colima", "DGO": "Durango", "GRO": "Guerrero",
    "GTO": "Guanajuato", "HGO": "Hidalgo", "JAL": "Jalisco", "MCH": "Michoacán",
    "MEX": "Estado de México", "MOR": "Morelos", "NAY": "Nayarit",
    "NL": "Nuevo León", "OAX": "Oaxaca", "PUE": "Puebla", "QR": "Quintana Roo",
    "QRO": "Querétaro", "SIN": "Sinaloa", "SLP": "San Luis Potosí",
    "SON": "Sonora", "TAB": "Tabasco", "TLX": "Tlaxcala", "TMS": "Tamaulipas",
    "VER": "Veracruz", "YUC": "Yucatán", "ZAC": "Zacatecas",
}

# Porcelanite nombra al distribuidor en minusculas y a veces solo con el
# apellido del dueno ("garcia", "kurodaak"). Estos son los que resultaron ser
# empresas que la BD ya tiene: 10 se reconocieron por nombre y 10 cruzando las
# coordenadas de sus sucursales contra las de la BD.
NOMBRE_COMERCIAL = {
    "Comercializadoras DMHC": "SODIMAC",
    "aquino": "MATERIAL PARA LA CONSTRUCCIÓN LA MISERICORDIA",
    "ceramat": "CERAMAT",
    "diaz": "ACABADOS CONTEMPORANEOS",
    "dival": "DIVAL SA DE CV",
    "ferretodo": "FERRETERO",
    "garcia": "GRUPO PLOMERÍA GARCÍA",
    "gersa": "GERSA",
    "gilsa": "GRUPO GILSA",
    "hinojosa": "EXPOCERAMICAS",   # sus tiendas: "Expoceramicas Av Mexico", Av. Ninos Heroes Gdl
    "kurodaak": "KURODA",
    "limsa": "LIMSA",
    "llano de la torre": "LLANO DE LA TORRE",
    "mcp": "MCP",
    "pereira": "PISOS Y RECUBRIMIENTOS",
    "recubre": "RECUBRE",
    "sanimex": "SANIMEX",
    "surtidor": "EL SURTIDOR",
    "villanueva": "LA CASA DEL BAÑO",
    "zetuna": "ZETUNA",
    "Leonel Garcia Fernandez": "FERREMAX",
    "Maria Estela Kurodasan": "KURODA",
    "PYM durango": "PISOS Y MÁS",
    "PYM queretaro": "PISOS Y MÁS",
    "arzuaga": "MEGA PISOS",
    # Confirmado contra gigantedelosazulejos.com: sus 45 sucursales estan en
    # Veracruz, Puebla, Tamaulipas, SLP, Oaxaca e Hidalgo, igual que estas.
    "cienfuegos": "EL GIGANTE DE LOS AZULEJOS Y MARMOLES",
    # Otro negocio, el de Guerrero y Morelos. La BD ya lo tiene por su nombre
    # comercial: su "Suc Galeana" y "EL GIGANTE DE LOS AZULEJOS DE JOJUTLA"
    # estan las dos en la Carretera Alpuyeca-Jojutla Km 13.
    "bernardo meneses": "EL GIGANTE DE LOS AZULEJOS",
    "dagda": "AZULEJARA SAN JOSE",
    "fierros y laminas": "EL BAZAR UNIVERSAL",
    "gpo garcia fernandez": "FERRESIN",
    "gpo quiroz ecohome": "ECOHOME",
    "4carrera": "EL 4 CARRERA PISOS Y BAÑOS SA DE CV",
    "Maderera y ferret. blanquitas ade": "CONSTRURAMA BLANQUITA",
    "PYM celaya": "PISOS Y MÁS",
    "PYM sanluis": "PISOS Y MÁS",
    "alce pisos y baños s. de r.l. de c.v.": "ALCE",
    "grupo azulejero de mayoristas norte s.a de c.v.": "SANIMEX",
    "kurodajak": "KURODA",
    "luigi": "EL CIRCO DEL MARMOL",
    "mahide": "DECORAPISOS SA DE CV",
    "materiales bucio y yañez s.a. de c.v.": "AZUMICH",
    "mirsa": "MIRSA TILE",
    "nevarez": "AZULEJOS LAS GARZAS SA DE CV",
    "su-kasa la falcon s.a. de c.v.": "SUKASA",
}


def sin_acentos(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(s or ""))
        if unicodedata.category(c) != "Mn"
    )


MENORES = {"de", "del", "la", "las", "los", "y", "el", "en", "a", "con"}


def titulo(s):
    """Title Case, como escriben los otros cuatro archivos de marcas.

    Porcelanite manda todo en mayusculas y sin espacio tras el punto
    ("BLVD.SALINAS"); se separa para que quede "Blvd. Salinas".
    """
    s = re.sub(r"\.(?=[A-Za-z\u00c0-\u017f])", ". ", str(s or "")).strip()
    if not s:
        return ""
    palabras = s.split()
    return " ".join(
        p.lower() if i > 0 and p.lower() in MENORES else p[:1].upper() + p[1:].lower()
        for i, p in enumerate(palabras)
    )


def mayus(s):
    """DISTRIBUIDOR va en mayusculas en los cinco archivos (2,418 de 2,419 en la BD)."""
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


# Los dos primeros digitos del codigo postal dicen el estado. Solo se usa
# cuando falta la abreviatura: se probo al reves y salio peor, porque el rango
# de codigos no distingue municipios que caen en el limite de dos estados.
CP_ESTADO = {}
for _ini, _fin, _edo in [
    (1, 16, "CDMX"), (20, 20, "Aguascalientes"), (21, 22, "Baja California"),
    (23, 23, "Baja California Sur"), (24, 24, "Campeche"), (25, 27, "Coahuila"),
    (28, 28, "Colima"), (29, 30, "Chiapas"), (31, 33, "Chihuahua"),
    (34, 35, "Durango"), (36, 38, "Guanajuato"), (39, 41, "Guerrero"),
    (42, 43, "Hidalgo"), (44, 49, "Jalisco"), (50, 57, "Estado de México"),
    (58, 61, "Michoacán"), (62, 62, "Morelos"), (63, 63, "Nayarit"),
    (64, 67, "Nuevo León"), (68, 71, "Oaxaca"), (72, 75, "Puebla"),
    (76, 76, "Querétaro"), (77, 77, "Quintana Roo"), (78, 79, "San Luis Potosí"),
    (80, 82, "Sinaloa"), (83, 85, "Sonora"), (86, 86, "Tabasco"),
    (87, 89, "Tamaulipas"), (90, 90, "Tlaxcala"), (91, 96, "Veracruz"),
    (97, 97, "Yucatán"), (98, 99, "Zacatecas"),
]:
    for _n in range(_ini, _fin + 1):
        CP_ESTADO[_n] = _edo

# Fuera de esta caja no hay territorio mexicano; un registro cae en Madrid.
CAJA_MX = (14.3, 32.8, -118.6, -86.5)


def estado_por_cp(cp):
    cp = re.sub(r"\D", "", str(cp or ""))
    if len(cp) != 5:
        return None
    return CP_ESTADO.get(int(cp[:2]))


def descargar():
    """El sitio filtra por huella TLS; curl_cffi imita la de Chrome."""
    try:
        from curl_cffi import requests
    except ImportError:
        sys.exit("Falta curl_cffi.  Instalalo con:  pip install curl_cffi")
    try:
        r = requests.get(URL, impersonate="chrome", timeout=60)
    except Exception as e:
        sys.exit(f"No se pudo descargar {URL}: {e}")
    if r.status_code != 200:
        sys.exit(f"{URL} respondio {r.status_code}")
    return r.text


def extraer(html):
    m = re.search(r"['\"]([A-Za-z0-9+/=]{5000,})['\"]", html)
    if not m:
        sys.exit("La pagina cambio: ya no trae el bloque base64 del mapa.")
    try:
        datos = json.loads(base64.b64decode(m.group(1)).decode("utf-8"))
    except Exception as e:
        sys.exit(f"No se pudo leer el bloque de datos: {e}")
    return datos.get("places", [])


def direccion(x):
    """El campo viene como "CALLE,NUMERO,CIUDAD,CP,PAIS"."""
    partes = [p.strip() for p in str(x.get("address") or "").split(",")]
    calle = partes[0] if partes else ""
    numero = partes[1] if len(partes) > 1 else ""
    cp = (x.get("location", {}) or {}).get("postal_code") or ""
    trozos = [titulo(" ".join(t for t in (calle, numero) if t))]
    if cp:
        trozos.append(f"CP {cp}")
    return ", ".join(t for t in trozos if t)


VIAS_DIR = {"av", "avenida", "blvd", "boulevard", "calle", "calz", "calzada",
            "carr", "carretera", "prol", "col", "colonia", "int", "num",
            "numero", "esq", "local", "lote", "sur", "norte", "ote", "pte",
            "centro", "fracc", "esquina"}


def sin_cp(direccion):
    """Quita el codigo postal para que no se confunda con el numero del local.

    Se busca por la etiqueta "CP" ademas de por largo: los codigos de la Ciudad
    de Mexico empiezan con cero y Excel se lo come, asi que 03300 llega escrito
    como 3300 y pasaria por numero exterior.
    """
    t = re.sub(r"[^\w\s]", " ", sin_acentos(str(direccion or "")).lower())
    t = re.sub(r"\bc\s*p\s*:?\s*\d{4,5}\b", " ", t)
    return " ".join(re.sub(r"\b\d{5}\b", " ", t).split())


def numero_exterior(direccion):
    nums = [int(n) for n in re.findall(r"\b(\d{1,5})\b", sin_cp(direccion))]
    return max(nums) if nums else None


def palabras_calle(direccion):
    return {w for w in re.sub(r"\d+", " ", sin_cp(direccion)).split()
            if len(w) > 3 and w not in VIAS_DIR}


def sin_repetidos(filas):
    """Quita la misma tienda repetida dentro del propio archivo.

    Porcelanite registra algunas dos veces con la direccion escrita distinto:
    "Paseo Cuauhnahuac Km 3.5" y "Boulevard Cuauhnahuac Km 3.5", o "Magnolias
    142" y "Av. Magnolias 142". Se comparan por distribuidor, numero exterior y
    cercania, no por texto.
    """
    por_dist = {}
    for f in filas:
        por_dist.setdefault(f[0], []).append(f)
    fuera, repetidas = set(), 0
    for grupo in por_dist.values():
        for i, a in enumerate(grupo):
            if id(a) in fuera:
                continue
            for b in grupo[i + 1:]:
                if id(b) in fuera:
                    continue
                d = math.hypot((a[6] - b[6]) * 110.57,
                               (a[7] - b[7]) * 111.32 * math.cos(math.radians(a[6])))
                if d > 0.3:
                    continue
                na_, nb = numero_exterior(a[5]), numero_exterior(b[5])
                if na_ is not None and na_ == nb and (palabras_calle(a[5]) & palabras_calle(b[5])):
                    fuera.add(id(b))
                    repetidas += 1
    if repetidas:
        print(f"  {repetidas} duplicados descartados")
    return [f for f in filas if id(f) not in fuera]


def avisar_si_hay_ediciones():
    if not DESTINO.exists():
        return
    generado = (MARCA_GENERADO.exists()
                and MARCA_GENERADO.read_text().strip() == str(int(DESTINO.stat().st_mtime)))
    if generado:
        return
    print(f"\n⚠️  {DESTINO.name} cambio desde la ultima vez que lo genero este script.")
    print("   Si lo editaste a mano, volver a bajar el sitio borrara esos cambios.\n")
    if input("   ¿Sobrescribir de todos modos? [s/N] ").strip().lower() not in ("s", "si", "sí"):
        sys.exit("   Cancelado: el Excel quedo intacto.")


def main():
    avisar_si_hay_ediciones()
    print(f"Descargando {URL} ...")
    lugares = extraer(descargar())
    print(f"  {len(lugares)} registros en la pagina")

    filas, sin_coords, fuera_mx = [], 0, 0
    for x in lugares:
        loc = x.get("location") or {}
        try:
            lat, lon = float(loc.get("lat")), float(loc.get("lng"))
        except (TypeError, ValueError):
            sin_coords += 1
            continue
        smin, smax, wmin, wmax = CAJA_MX
        if not (smin <= lat <= smax and wmin <= lon <= wmax):
            fuera_mx += 1
            continue

        cats = x.get("categories") or []
        crudo = cats[0]["name"].strip() if cats else ""
        dist = NOMBRE_COMERCIAL.get(crudo, mayus(crudo))
        if not dist:
            continue

        abrev = (loc.get("state") or "").strip().upper()
        estado = (ESTADOS.get(abrev) or estado_por_cp(loc.get("postal_code"))
                  or titulo(abrev))
        filas.append([
            dist,
            "México",
            titulo(x.get("title")),
            estado,
            titulo(loc.get("city")),
            direccion(x),
            round(lat, 6),
            round(lon, 6),
            f"https://www.google.com/maps?q={lat},{lon}",
        ])

    print(f"  {len(filas)} con coordenadas")
    if sin_coords:
        print(f"  {sin_coords} sin coordenadas (no se pueden mapear)")
    if fuera_mx:
        print(f"  {fuera_mx} con coordenada fuera de Mexico (descartados)")

    filas = sin_repetidos(filas)

    filas.sort(key=lambda f: (sin_acentos(f[0]).upper(), sin_acentos(f[2]).upper()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sucursales Porcelanite"
    ws.append(COLUMNAS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for f in filas:
        ws.append(f)
    for i, ancho in enumerate([46, 10, 34, 20, 26, 62, 12, 12, 52], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DESTINO)
    MARCA_GENERADO.write_text(str(int(DESTINO.stat().st_mtime)))

    print(f"\nListo: {DESTINO}")
    print(f"  distribuidores: {len({f[0] for f in filas})}")


if __name__ == "__main__":
    main()
