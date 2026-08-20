#!/usr/bin/env python3
"""Completa la BD con sucursales que las marcas conocen y el padron no tiene.

Uso:
    python3 altas_faltantes.py            # muestra que agregaria
    python3 altas_faltantes.py --aplicar  # escribe en Distribuidores.xlsx

No es lo mismo que `altas_porcelanite.py` o `altas_urrea.py`. Aquellos daban de
alta empresas nuevas; este solo rellena las tiendas que le faltan a empresas que
la BD ya tiene, con el nombre comercial que la BD ya usa. Asi no entra ninguna
razon social.

De donde sale el hueco: cada marca publica su propia red y ninguna coincide con
el padron. De GRUPO VAMA la BD tiene 4 sucursales y las marcas reportan 40; de
GRUPO BOXITO, 58 contra 170.

Solo se tocan los distribuidores con MINIMO_TIENDAS o mas sucursales en la BD.
En una cadena grande, que la marca reporte una tienda mas es casi seguro una
sucursal que faltaba. En una de dos tiendas, es mas probable que sea otra
empresa con nombre parecido.
"""

import math
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from limpiar_bd import misma_tienda, na

SERVIDOR = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales"
)
BD = SERVIDOR / "Distribuidores.xlsx"
# (archivo, columnas: nombre, estado, ciudad, direccion, latitud)
MARCAS = [
    ("Marcas/Daltile.xlsx", 2, 3, 4, 5, 6),
    ("Marcas/Cesantoni.xlsx", 2, 3, 4, 5, 6),
    ("Marcas/Vitromex.xlsx", 2, 3, 4, 5, 6),
    ("Marcas/Porcelanite.xlsx", 2, 3, 4, 5, 6),
    ("Marcas/Urrea.xlsx", 2, 3, 4, 5, 6),
]
# Interceramic queda fuera a proposito: son sus propias tiendas, no distribuidores.

MINIMO_TIENDAS = 10
LEJOS_KM = 30   # mas alla de esto ni se comparan; son ciudades distintas


def km(a, b, c, d):
    return (((a - c) * 110.57) ** 2
            + ((b - d) * 111.32 * math.cos(math.radians(a))) ** 2) ** 0.5


def main():
    aplicar = "--aplicar" in sys.argv

    libro = openpyxl.load_workbook(BD)   # sin data_only: conserva las formulas
    hoja = libro["BD"]
    tiendas = defaultdict(list)
    for fila in hoja.iter_rows(min_row=2):
        if not fila[0].value:
            continue
        nombre = " ".join(str(fila[0].value).split())
        try:
            lat, lon = float(fila[7].value), float(fila[8].value)
        except (TypeError, ValueError):
            continue
        tiendas[nombre].append(
            (str(fila[6].value or ""), str(fila[5].value or ""),
             str(fila[4].value or ""), lat, lon)
        )

    grandes = {n for n, v in tiendas.items() if len(v) >= MINIMO_TIENDAS}
    print(f"  distribuidores en la BD: {len(tiendas)}")
    print(f"  con {MINIMO_TIENDAS} o mas sucursales: {len(grandes)}\n")

    faltantes, de_marca = [], Counter()
    for archivo, cN, cE, cC, cA, cLA in MARCAS:
        ruta = SERVIDOR / archivo
        if not ruta.exists():
            continue
        marca = archivo.split("/")[-1][:-5]
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        for f in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True):
            if not f[0]:
                continue
            nombre = " ".join(str(f[0]).split())
            if nombre not in grandes:
                continue
            try:
                lat, lon = float(f[cLA]), float(f[cLA + 1])
            except (TypeError, ValueError):
                continue
            señas = (str(f[cA] or ""), str(f[cC] or ""), str(f[cE] or ""))
            ya = any(
                misma_tienda(señas, (d, c, e), km(lat, lon, x, y) * 1000)
                for d, c, e, x, y in tiendas[nombre]
                if km(lat, lon, x, y) < LEJOS_KM
            )
            if ya:
                continue
            faltantes.append([nombre, str(f[cN] or ""), str(f[cE] or ""),
                              str(f[cC] or ""), str(f[cA] or ""), lat, lon])
            # se apunta enseguida para que otra marca no la proponga otra vez
            tiendas[nombre].append((str(f[cA] or ""), str(f[cC] or ""),
                                    str(f[cE] or ""), lat, lon))
            de_marca[marca] += 1
        wb.close()

    print(f"  ALTAS: {len(faltantes)} sucursales de {len({f[0] for f in faltantes})} distribuidores")
    print(f"  por marca: {dict(de_marca)}\n")
    for nombre, n in Counter(f[0] for f in faltantes).most_common(20):
        print(f"     {n:>4}  {nombre[:40]:42} (la BD tenia {len(tiendas[nombre]) - n})")

    if not aplicar:
        print("\n  Nada se escribio. Para aplicarlo:  python3 altas_faltantes.py --aplicar")
        return

    n = hoja.max_row
    for dist, nom, est, ciu, dire, lat, lon in faltantes:
        n += 1
        hoja.cell(n, 1, dist)
        hoja.cell(n, 2, nom)
        hoja.cell(n, 4, "México")
        hoja.cell(n, 5, est)
        hoja.cell(n, 6, ciu)
        hoja.cell(n, 7, dire)
        hoja.cell(n, 8, lat)
        hoja.cell(n, 9, lon)
        hoja.cell(n, 10, f'=IF(H{n}="","","https://www.google.com/maps?q="&H{n}&","&I{n})')
    libro.save(BD)
    print(f"\n  BD queda en {hoja.max_row - 1} filas")


if __name__ == "__main__":
    main()
