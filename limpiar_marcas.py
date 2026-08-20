#!/usr/bin/env python3
"""Revisa los Excel de las marcas: duplicados y nombres contra la BD.

Uso:
    python3 limpiar_marcas.py            # revisa y reporta
    python3 limpiar_marcas.py --aplicar  # borra los duplicados

Cada archivo de marca es la red que esa marca publica, y llega con los defectos
de quien la capturo: la misma tienda dos veces, o el distribuidor escrito de una
forma que no coincide con la BD.

Se hacen dos cosas:

* **Duplicados dentro del archivo.** Misma tienda, mismo distribuidor, dos filas.
  Se compara con el mismo criterio que la BD, que ya entiende las direcciones
  por manzana y lote de la peninsula y las calles numeradas de Merida.
* **Nombres que la BD no conoce.** No se cambian solos: se listan para
  revisarlos, porque un nombre que no esta en la BD puede ser un distribuidor
  que de verdad falta o el mismo escrito de otra manera.

Interceramic entra en la revision de duplicados pero no en la de nombres: son
sus propias tiendas, no distribuidores, y no tiene por que estar en la BD.
"""

import math
import re
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from limpiar_bd import misma_tienda, riqueza

SERVIDOR = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales"
)
BD = SERVIDOR / "Distribuidores.xlsx"
MARCAS = ["Daltile", "Interceramic", "Cesantoni", "Vitromex", "Porcelanite", "Urrea"]
APARTE = {"Interceramic"}   # sus tiendas propias, no distribuidores

# columnas en los archivos de marca: distribuidor, nombre, estado, ciudad,
# direccion, latitud (la longitud va enseguida)
COL = dict(dist=0, nombre=2, estado=3, ciudad=4, direccion=5, lat=6)


def na(s):
    s = "".join(c for c in unicodedata.normalize("NFD", str(s or "").lower())
                if unicodedata.category(c) != "Mn")
    return " ".join(re.sub(r"[^\w\s]", " ", s).split())


def km(a, b, c, d):
    return (((a - c) * 110.57) ** 2
            + ((b - d) * 111.32 * math.cos(math.radians(a))) ** 2) ** 0.5


def nombres_de_la_bd():
    wb = openpyxl.load_workbook(BD, read_only=True, data_only=True)
    fuera = {na(f[0]) for f in wb["BD"].iter_rows(min_row=2, values_only=True) if f[0]}
    wb.close()
    return fuera


def main():
    aplicar = "--aplicar" in sys.argv
    en_bd = nombres_de_la_bd()
    print(f"  la BD conoce {len(en_bd)} distribuidores\n")

    resumen = []
    for marca in MARCAS:
        ruta = SERVIDOR / "Marcas" / f"{marca}.xlsx"
        if not ruta.exists():
            continue
        libro = openpyxl.load_workbook(ruta)
        hoja = libro[libro.sheetnames[0]]

        filas = []
        for fila in hoja.iter_rows(min_row=2):
            if not fila[COL["dist"]].value:
                continue
            try:
                lat = float(fila[COL["lat"]].value)
                lon = float(fila[COL["lat"] + 1].value)
            except (TypeError, ValueError):
                continue
            filas.append((fila, " ".join(str(fila[COL["dist"]].value).split()), lat, lon))

        por_dist = defaultdict(list)
        for f in filas:
            por_dist[f[1]].append(f)

        def señas(f):
            return (str(f[0][COL["direccion"]].value or ""),
                    str(f[0][COL["ciudad"]].value or ""),
                    str(f[0][COL["estado"]].value or ""))

        fuera, pares = set(), []
        for dist, grupo in por_dist.items():
            for i in range(len(grupo)):
                a = grupo[i]
                if a[0][0].row in fuera:
                    continue
                for b in grupo[i + 1:]:
                    if b[0][0].row in fuera:
                        continue
                    metros = km(a[2], a[3], b[2], b[3]) * 1000
                    if not misma_tienda(señas(a), señas(b), metros):
                        continue
                    queda, sobra = ((a, b) if riqueza(a[0]) >= riqueza(b[0]) else (b, a))
                    fuera.add(sobra[0][0].row)
                    pares.append((dist, queda, sobra))

        desconocidos = Counter()
        if marca not in APARTE:
            for dist, grupo in por_dist.items():
                if na(dist) not in en_bd:
                    desconocidos[dist] = len(grupo)

        print(f"=== {marca}: {len(filas)} filas, {len(por_dist)} distribuidores ===")
        print(f"   duplicados: {len(pares)}")
        for dist, cuenta in Counter(p[0] for p in pares).most_common(5):
            print(f"      {cuenta:>3}  {dist[:44]}")
        if marca in APARTE:
            print("   nombres: no aplica, son tiendas propias de la marca")
        else:
            print(f"   distribuidores que la BD no tiene: {len(desconocidos)}"
                  f"  ({sum(desconocidos.values())} sucursales)")
            for dist, cuenta in desconocidos.most_common(8):
                print(f"      {cuenta:>4}  {dist[:46]}")
        print()
        resumen.append((marca, len(filas), len(pares), len(desconocidos)))

        if aplicar and fuera:
            for r in sorted(fuera, reverse=True):
                hoja.delete_rows(r)
            for i in range(6):
                try:
                    libro.save(ruta)
                    break
                except OSError:
                    if i == 5:
                        raise
                    time.sleep(4)

    print("  resumen:")
    for marca, filas, dup, desc in resumen:
        print(f"     {marca:14} {filas:>5} filas   {dup:>4} duplicados   {desc:>4} nombres fuera de la BD")
    if not aplicar:
        print("\n  Nada se escribio. Para aplicarlo:  python3 limpiar_marcas.py --aplicar")


if __name__ == "__main__":
    main()
