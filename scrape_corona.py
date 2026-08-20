#!/usr/bin/env python3
"""Genera Marcas/Corona.xlsx desde el buscador de tiendas de Corona.

Uso:
    python3 scrape_corona.py

https://coronamexico.com/store-finder corre sobre SAP Hybris y pide los puntos
de venta a `/store-finder/position?latitude=&longitude=&page=`. Contesta de diez
en diez ordenados por cercania, pero el campo `total` dice cuantos hay en toda
la red, y paginando desde un solo punto salen los 989.

Lo que no trae: estado, y casi nunca ciudad. Se sacan del resto de los
archivos, que ya tienen 8,315 puntos con su municipio: 830 de las 989 tiendas
de Corona caen a menos de un kilometro de uno conocido.

Tampoco separa al distribuidor de la sucursal: el mismo campo dice "Sanimex
(Sanimex Ayuntamiento)", "MARRODAN SUC. ALLENDE" o simplemente "AYUNTAMIENTO
14". Se identifica contra la BD por nombre, por prefijo y por direccion.
"""

import json
import math
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from limpiar_bd import misma_tienda

URL = "https://coronamexico.com/store-finder/position"
PAGINA = "https://coronamexico.com/store-finder"
AQUI = Path(__file__).resolve().parent
SERVIDOR = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales"
)
DESTINO = SERVIDOR / "Marcas" / "Corona.xlsx"
MARCA_GENERADO = AQUI / ".corona-generado"

COLUMNAS = [
    "DISTRIBUIDOR", "PAÍS", "NOMBRE", "ESTADO", "CIUDAD",
    "DIRECCIÓN", "LATITUD", "LONGITUD", "URL DE UBICACIÓN",
]

# Los demas archivos, para sacar municipio y estado por cercania y para
# reconocer al distribuidor.
GAZETTEER = [
    ("Distribuidores.xlsx", "BD", 4, 5, 7),
    ("Marcas/Daltile.xlsx", None, 3, 4, 6),
    ("Marcas/Cesantoni.xlsx", None, 3, 4, 6),
    ("Marcas/Vitromex.xlsx", None, 3, 4, 6),
    ("Marcas/Porcelanite.xlsx", None, 3, 4, 6),
    ("Marcas/Urrea.xlsx", None, 3, 4, 6),
]

MENORES = {"de", "del", "la", "las", "los", "y", "el", "en", "a", "con"}
GENERICAS = {"suc", "sucursal", "matriz", "bodega", "cedis", "tienda"}
CERCA_CIUDAD_KM = 15

# Fuera de esta caja no hay territorio mexicano. Corona manda 35 tiendas con la
# coordenada podrida: "Gersa Hermosillo" cae en Carolina del Sur, "Kuroda Son
# Obregón" en Zaragoza, y "Ceramat Tapachula" trae la latitud repetida como
# longitud. Todas dicen ciudad mexicana en el nombre, asi que no son sucursales
# en el extranjero: es el dato que viene mal.
CAJA_MX = (14.3, 32.8, -118.6, -86.5)


def sin_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                   if unicodedata.category(c) != "Mn")


def na(s):
    return " ".join(re.sub(r"[^\w\s]", " ", sin_acentos(s).lower()).split())


def titulo(s):
    s = re.sub(r"\.(?=[A-Za-zÀ-ſ])", ". ", str(s or "")).strip()
    if not s:
        return ""
    return " ".join(
        p.lower() if i > 0 and p.lower() in MENORES else p[:1].upper() + p[1:].lower()
        for i, p in enumerate(s.split())
    )


def km(a, b, c, d):
    return (((a - c) * 110.57) ** 2
            + ((b - d) * 111.32 * math.cos(math.radians(a))) ** 2) ** 0.5


def bajar():
    """Pagina el buscador hasta que deja de contestar."""
    try:
        from curl_cffi import requests
    except ImportError:
        sys.exit("Falta curl_cffi.  Instalalo con:  pip install curl_cffi")
    s = requests.Session()
    s.get(PAGINA, impersonate="chrome", timeout=45, verify=False)
    tiendas, pagina = {}, 0
    while pagina < 200:
        try:
            r = s.get(URL, params=dict(latitude=19.43, longitude=-99.13, page=pagina),
                      impersonate="chrome", timeout=45, verify=False,
                      headers={"Accept": "application/json",
                               "X-Requested-With": "XMLHttpRequest"})
            datos = r.json().get("data") or []
        except Exception:
            break
        if not datos:
            break
        for x in datos:
            tiendas[x["name"]] = x
        pagina += 1
    if not tiendas:
        sys.exit("El buscador no contesto; puede que haya cambiado el endpoint.")
    return list(tiendas.values())


def cargar_referencia():
    """Puntos conocidos, y los nombres de distribuidor de la BD."""
    puntos, en_bd = [], []
    for archivo, hoja, cE, cC, cLA in GAZETTEER:
        ruta = SERVIDOR / archivo
        if not ruta.exists():
            continue
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]
        cA = 6 if hoja == "BD" else 5
        for f in ws.iter_rows(min_row=2, values_only=True):
            if not f[0]:
                continue
            try:
                lat, lon = float(f[cLA]), float(f[cLA + 1])
            except (TypeError, ValueError):
                continue
            nombre = " ".join(str(f[0]).split())
            puntos.append((lat, lon, str(f[cE] or ""), str(f[cC] or "")))
            if hoja == "BD":
                en_bd.append((nombre, str(f[cA] or ""), str(f[cC] or ""),
                              str(f[cE] or ""), lat, lon))
        wb.close()
    return puntos, en_bd


def indice(filas, llave, decimales=2):
    """Reparte los puntos en celdas para no comparar todos contra todos.

    El tamano de celda tiene que ir con el radio que se va a buscar: con dos
    decimales la celda es de un kilometro, y buscar a saltos de 0.1 grados se
    salta todo lo que hay en medio.
    """
    cubo = defaultdict(list)
    for r in filas:
        lat, lon = llave(r)
        cubo[(round(lat, decimales), round(lon, decimales))].append(r)
    return cubo


def vecinos(cubo, lat, lon, decimales=2):
    paso = 10 ** -decimales
    salida = []
    for dx in (-paso, 0, paso):
        for dy in (-paso, 0, paso):
            salida += cubo.get((round(lat + dx, decimales), round(lon + dy, decimales)), [])
    return salida


def tokens(nombre):
    return [w for w in na(nombre).split() if w not in MENORES and w not in GENERICAS]


def identificar(display, direccion, lat, lon, en_bd, por_nombre, raros, cubo_bd):
    """A que distribuidor de la BD pertenece esta tienda, si a alguno."""
    d = na(display)

    # 1. El nombre de la BD tal cual, al principio
    mejor = None
    for nombre, tk in por_nombre:
        if d == tk or d.startswith(tk + " "):
            if mejor is None or len(tk) > len(mejor[1]):
                mejor = (nombre, tk)
    if mejor:
        return mejor[0]

    # 2. La direccion coincide con una tienda de la BD
    for nombre, dire, ciu, est, x, y in vecinos(cubo_bd, lat, lon):
        metros = km(lat, lon, x, y) * 1000
        if metros < 400 and misma_tienda((direccion, "", ""), (dire, ciu, est), metros):
            return nombre

    # 3. Empieza con palabras suficientes de un nombre de la BD. Con una sola
    #    palabra se pide que sea rara, para no juntar cualquier "GRUPO" o
    #    "MATERIALES" con el primero que se le parezca.
    tk = tokens(display)
    for nombre, suyos in por_nombre_tokens(en_bd):
        n = len(suyos)
        if n and tk[:n] == suyos and (n >= 2 or suyos[0] in raros):
            return nombre
    return None


_cache_tokens = {}


def por_nombre_tokens(en_bd):
    if not _cache_tokens:
        vistos = {}
        for nombre, *_ in en_bd:
            vistos.setdefault(nombre, tokens(nombre))
        _cache_tokens.update(vistos)
    return sorted(_cache_tokens.items(), key=lambda kv: -len(kv[1]))


def avisar_si_hay_ediciones():
    if not DESTINO.exists():
        return
    if (MARCA_GENERADO.exists()
            and MARCA_GENERADO.read_text().strip() == str(int(DESTINO.stat().st_mtime))):
        return
    print(f"\n⚠️  {DESTINO.name} cambio desde la ultima vez que lo genero este script.\n")
    if input("   ¿Sobrescribir de todos modos? [s/N] ").strip().lower() not in ("s", "si", "sí"):
        sys.exit("   Cancelado: el Excel quedo intacto.")


def main():
    avisar_si_hay_ediciones()
    print(f"Consultando {URL} ...")
    tiendas = bajar()
    print(f"  {len(tiendas)} puntos de venta")

    puntos, en_bd = cargar_referencia()
    print(f"  referencia: {len(puntos)} puntos conocidos, {len({n for n, *_ in en_bd})} distribuidores en la BD")

    # una decima de grado son unos 11 km, que es el radio con que se busca
    # el municipio mas cercano
    cubo_geo = indice(puntos, lambda r: (r[0], r[1]), 1)
    cubo_bd = indice(en_bd, lambda r: (r[4], r[5]))
    por_nombre = sorted({(n, na(n)) for n, *_ in en_bd}, key=lambda kv: -len(kv[1]))
    frec = defaultdict(int)
    for nombre in {n for n, *_ in en_bd}:
        for w in set(tokens(nombre)):
            frec[w] += 1
    raros = {w for w, c in frec.items() if c == 1}

    filas, reconocidos, sin_ciudad, fuera_mx = [], 0, 0, []
    for x in tiendas:
        try:
            lat, lon = float(x["latitude"]), float(x["longitude"])
        except (TypeError, ValueError):
            continue
        smin, smax, wmin, wmax = CAJA_MX
        if not (smin <= lat <= smax and wmin <= lon <= wmax):
            fuera_mx.append((str(x.get("displayName") or ""), lat, lon))
            continue
        cp = re.sub(r"\D", "", str(x.get("postalCode") or ""))
        calle = titulo(x.get("line1"))
        direccion = ", ".join(t for t in (calle, f"CP {cp}" if cp else "") if t)

        estado, ciudad, cerca = "", titulo(x.get("town")), 9e9
        for a, b, e, c in vecinos(cubo_geo, lat, lon, 1):
            d = km(lat, lon, a, b)
            if d < cerca:
                cerca, estado = d, e
                if not ciudad:
                    ciudad = c
        if cerca > CERCA_CIUDAD_KM:
            estado = estado if cerca < 60 else ""
            sin_ciudad += 1

        display = str(x.get("displayName") or "").strip()
        dist = identificar(display, direccion, lat, lon, en_bd, por_nombre, raros, cubo_bd)
        if dist:
            reconocidos += 1
        else:
            dist = re.sub(r"\s+", " ", display).strip().upper()

        filas.append([
            dist, "México", titulo(display), estado, ciudad, direccion,
            round(lat, 6), round(lon, 6),
            f"https://www.google.com/maps?q={lat},{lon}",
        ])

    print(f"  distribuidor reconocido contra la BD: {reconocidos} de {len(filas)}")
    if sin_ciudad:
        print(f"  sin punto conocido a {CERCA_CIUDAD_KM} km: {sin_ciudad}")
    if fuera_mx:
        print(f"  {len(fuera_mx)} con coordenada fuera de Mexico (descartados):")
        for nombre, la, lo in fuera_mx[:6]:
            print(f"      {nombre[:38]:40} {la},{lo}")

    vistas, unicas = set(), []
    for f in filas:
        k = (na(f[0]), round(f[6], 5), round(f[7], 5))
        if k in vistas:
            continue
        vistas.add(k)
        unicas.append(f)
    if len(unicas) < len(filas):
        print(f"  {len(filas) - len(unicas)} duplicados descartados")
    filas = sorted(unicas, key=lambda f: (sin_acentos(f[0]).upper(), sin_acentos(f[4]).upper()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sucursales Corona"
    ws.append(COLUMNAS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for f in filas:
        ws.append(f)
    for i, ancho in enumerate([46, 10, 34, 20, 26, 62, 12, 12, 52], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DESTINO)
    MARCA_GENERADO.write_text(str(int(DESTINO.stat().st_mtime)))
    print(f"\nListo: {DESTINO}")
    print(f"  filas: {len(filas)}   distribuidores: {len({f[0] for f in filas})}")


if __name__ == "__main__":
    main()
