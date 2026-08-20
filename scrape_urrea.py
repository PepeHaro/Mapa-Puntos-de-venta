#!/usr/bin/env python3
"""Genera Marcas/Urrea.xlsx desde el localizador de Urrea.

Uso:
    python3 scrape_urrea.py

El buscador de https://urrea.mx/ pide los distribuidores a la API
`POST /api/public/locate/dealers` mandando {lat, lng}, y contesta solo los que
caen en un radio de 5 km. Barrer el pais asi serian unas 40,000 peticiones,
pero si el cuerpo va vacio la API devuelve los 3,414 de una sola vez.

No se guarda el telefono que trae cada registro: el mapa es publico.
"""

import json
import math
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

URL = "https://urrea.mx/api/public/locate/dealers"
AQUI = Path(__file__).resolve().parent
DESTINO = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales/Marcas/Urrea.xlsx"
)
MARCA_GENERADO = AQUI / ".urrea-generado"

COLUMNAS = [
    "DISTRIBUIDOR", "PAÍS", "NOMBRE", "ESTADO", "CIUDAD",
    "DIRECCIÓN", "LATITUD", "LONGITUD", "URL DE UBICACIÓN",
]

# Urrea escribe el estado de 47 maneras distintas: con y sin acento, en
# mayusculas, con el nombre largo oficial y con erratas ("Ciudadd de México").
ESTADOS = {
    "aguascalientes": "Aguascalientes",
    "baja california": "Baja California",
    "baja california sur": "Baja California Sur",
    "campeche": "Campeche",
    "chiapas": "Chiapas",
    "chihuahua": "Chihuahua",
    "ciudad de mexico": "CDMX",
    "ciudadd de mexico": "CDMX",
    "coahuila de zaragoza": "Coahuila",
    "colima": "Colima",
    "durango": "Durango",
    "estado de mexico": "Estado de México",
    "mexico": "Estado de México",
    "guanajuato": "Guanajuato",
    "guerrero": "Guerrero",
    "hidalgo": "Hidalgo",
    "jalisco": "Jalisco",
    "michoacan": "Michoacán",
    "michoacan de ocampo": "Michoacán",
    "morelos": "Morelos",
    "nayarit": "Nayarit",
    "nuevo leon": "Nuevo León",
    "oaxaca": "Oaxaca",
    "puebla": "Puebla",
    "queretaro": "Querétaro",
    "quintana roo": "Quintana Roo",
    "san luis potosi": "San Luis Potosí",
    "sinaloa": "Sinaloa",
    "sonora": "Sonora",
    "tabasco": "Tabasco",
    "tamaulipas": "Tamaulipas",
    "tlaxcala": "Tlaxcala",
    "veracruz de ignacio de la llave": "Veracruz",
    "veracruz": "Veracruz",
    "yucatan": "Yucatán",
    "zacatecas": "Zacatecas",
}

MENORES = {"de", "del", "la", "las", "los", "y", "el", "en", "a", "con"}

VIAS_DIR = {"av", "avenida", "blvd", "boulevard", "calle", "calz", "calzada",
            "carr", "carretera", "prol", "col", "colonia", "int", "num",
            "numero", "esq", "local", "lote", "sur", "norte", "ote", "pte",
            "centro", "fracc", "esquina"}

# Coordenadas que Urrea manda mal y se corrigen a mano. La llave es el id del
# distribuidor en su API.
COORDENADAS_A_MANO = {}

# Fuera de esta caja no hay territorio mexicano. Urrea usa dos rellenos para
# "no se donde esta": 0,0 y 1,-1, los dos frente a Africa.
CAJA_MX = (14.3, 32.8, -118.6, -86.5)


def sin_acentos(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(s or ""))
        if unicodedata.category(c) != "Mn"
    )


def titulo(s):
    """Title Case, como escriben los otros archivos de marcas."""
    s = re.sub(r"\.(?=[A-Za-zÀ-ſ])", ". ", str(s or "")).strip()
    if not s:
        return ""
    return " ".join(
        p.lower() if i > 0 and p.lower() in MENORES else p[:1].upper() + p[1:].lower()
        for i, p in enumerate(s.split())
    )


def sin_razon_social(nombre):
    """Quita las siglas de la forma legal del final del nombre.

    Urrea escribe el mismo negocio de varias maneras y cada una cuenta como
    distribuidor aparte: "ZETUNA", "ZETUNA S DE RL DE C.V." y "ZETUNA S DE RL
    DE CV" son tres. Al quitar las siglas los tres se vuelven "ZETUNA", que
    ademas es como lo tiene la BD.
    """
    t = " ".join(re.sub(r"[.,]", " ", str(nombre or "")).split())
    t = re.sub(r"\s+S\s*A\s*P\s*I(\s+DE)?(\s+C\s*V)?\s*$", "", t, flags=re.I)
    t = re.sub(r"\s+S(\s+DE)?\s+R\s*L(\s+DE)?(\s+C\s*V)?\s*$", "", t, flags=re.I)
    t = re.sub(r"\s+S\s*A(\s+DE?)?(\s+C\s*V?)?\s*$", "", t, flags=re.I)
    t = re.sub(r"\s+DE\s+C\s*V\s*$", "", t, flags=re.I)
    return t.strip() or " ".join(str(nombre or "").split())


def mayus(s):
    """DISTRIBUIDOR va en mayusculas en todos los archivos."""
    return re.sub(r"\s+", " ", sin_razon_social(s)).strip().upper()


def sin_cp(direccion):
    t = re.sub(r"[^\w\s]", " ", sin_acentos(str(direccion or "")).lower())
    t = re.sub(r"\bc\s*p\s*:?\s*\d{4,5}\b", " ", t)
    return " ".join(re.sub(r"\b\d{5}\b", " ", t).split())


def numero_exterior(direccion):
    nums = [int(n) for n in re.findall(r"\b(\d{1,5})\b", sin_cp(direccion))]
    return max(nums) if nums else None


def palabras_calle(direccion):
    return {w for w in re.sub(r"\d+", " ", sin_cp(direccion)).split()
            if len(w) > 3 and w not in VIAS_DIR}


def bajar():
    """El cuerpo vacio hace que la API conteste el padron completo."""
    try:
        from curl_cffi import requests
    except ImportError:
        sys.exit("Falta curl_cffi.  Instalalo con:  pip install curl_cffi")
    try:
        s = requests.Session()
        s.get("https://urrea.mx/", impersonate="chrome", timeout=45, verify=False)
        r = s.post(URL, impersonate="chrome", timeout=180, verify=False, json={},
                   headers={"Accept": "application/json", "Referer": "https://urrea.mx/"})
    except Exception as e:
        sys.exit(f"No se pudo consultar {URL}: {e}")
    if r.status_code != 200:
        sys.exit(f"{URL} respondio {r.status_code}")
    try:
        datos = r.json()
    except Exception as e:
        sys.exit(f"La respuesta no es JSON: {e}")
    if not isinstance(datos, list) or not datos:
        sys.exit("La API contesto vacio; puede que haya cambiado el contrato.")
    return datos


def sin_repetidos(filas):
    """Quita la misma tienda repetida dentro del propio padron."""
    por_dist = {}
    for f in filas:
        por_dist.setdefault(f[0], []).append(f)
    fuera, repetidas = set(), 0
    for grupo in por_dist.values():
        for i, a in enumerate(grupo):
            if id(a) in fuera:
                continue
            for b in grupo[i + 1:]:
                if id(b) in fuera:
                    continue
                d = math.hypot((a[6] - b[6]) * 110.57,
                               (a[7] - b[7]) * 111.32 * math.cos(math.radians(a[6])))
                if d > 0.3:
                    continue
                na_, nb = numero_exterior(a[5]), numero_exterior(b[5])
                if na_ is not None and na_ == nb and (palabras_calle(a[5]) & palabras_calle(b[5])):
                    fuera.add(id(b))
                    repetidas += 1
    if repetidas:
        print(f"  {repetidas} duplicados descartados")
    return [f for f in filas if id(f) not in fuera]


# "Muebles para Baño" no es una empresa: es el rotulo que usan las tiendas del
# grupo GERSA, y Urrea lo pone como si fuera el nombre del distribuidor en casi
# 300 sucursales. Las que se pueden confirmar por direccion contra la BD pasan a
# llamarse GERSA; las demas no se pueden identificar y no entran al archivo.
GENERICOS = {"MUEBLES PARA BAÑO", "MUEBLES PARA BAÑOS"}


def resolver_genericos(filas):
    """Renombra a GERSA las tiendas genericas cuya direccion ya esta en la BD."""
    bd = Path(str(DESTINO.parent.parent / "Distribuidores.xlsx"))
    if not bd.exists():
        return filas
    libro = openpyxl.load_workbook(bd, read_only=True, data_only=True)
    puntos = {}
    for f in libro["BD"].iter_rows(min_row=2, values_only=True):
        if not f[0]:
            continue
        try:
            la, lo = float(f[7]), float(f[8])
        except (TypeError, ValueError):
            continue
        puntos.setdefault((round(la, 2), round(lo, 2)), []).append(
            (la, lo, str(f[6] or ""), str(f[0]))
        )
    libro.close()

    resueltas, perdidas, salida = 0, 0, []
    for f in filas:
        if f[0] not in GENERICOS:
            salida.append(f)
            continue
        lat, lon = f[6], f[7]
        dueño = None
        for dla in (-0.01, 0, 0.01):
            for dlo in (-0.01, 0, 0.01):
                for x, y, direccion, nombre in puntos.get((round(lat + dla, 2), round(lon + dlo, 2)), []):
                    d = math.hypot((lat - x) * 110.57,
                                   (lon - y) * 111.32 * math.cos(math.radians(lat)))
                    if d > 0.15:
                        continue
                    na_, nb = numero_exterior(f[5]), numero_exterior(direccion)
                    if na_ is not None and na_ == nb and (palabras_calle(f[5]) & palabras_calle(direccion)):
                        dueño = nombre
        if dueño:
            f[0] = dueño
            f[2] = titulo(dueño)
            resueltas += 1
            salida.append(f)
        else:
            perdidas += 1
    if resueltas or perdidas:
        print(f"  \"Muebles para Baño\": {resueltas} identificadas por direccion, "
              f"{perdidas} sin identificar (no entran)")
    return salida


def avisar_si_hay_ediciones():
    if not DESTINO.exists():
        return
    generado = (MARCA_GENERADO.exists()
                and MARCA_GENERADO.read_text().strip() == str(int(DESTINO.stat().st_mtime)))
    if generado:
        return
    print(f"\n⚠️  {DESTINO.name} cambio desde la ultima vez que lo genero este script.")
    print("   Si lo editaste a mano, volver a bajarlo borrara esos cambios.\n")
    if input("   ¿Sobrescribir de todos modos? [s/N] ").strip().lower() not in ("s", "si", "sí"):
        sys.exit("   Cancelado: el Excel quedo intacto.")


def main():
    avisar_si_hay_ediciones()
    print(f"Consultando {URL} ...")
    padron = bajar()
    print(f"  {len(padron)} distribuidores en el padron")

    filas, sin_coords, fuera_mx = [], 0, 0
    for x in padron:
        lat, lng = COORDENADAS_A_MANO.get(x.get("id"), (x.get("lat"), x.get("lng")))
        try:
            lat, lng = float(lat), float(lng)
        except (TypeError, ValueError):
            sin_coords += 1
            continue
        if (lat, lng) in ((0, 0), (1, -1)):
            sin_coords += 1
            continue
        smin, smax, wmin, wmax = CAJA_MX
        if not (smin <= lat <= smax and wmin <= lng <= wmax):
            fuera_mx += 1
            continue

        estado = ESTADOS.get(sin_acentos(str(x.get("state") or "")).lower().strip(),
                             titulo(x.get("state")))
        cp = str(x.get("postal_code") or "").strip()
        direccion = titulo(x.get("address"))
        if cp:
            direccion = f"{direccion}, CP {cp}" if direccion else f"CP {cp}"

        filas.append([
            mayus(x.get("name")),
            "México",
            titulo(sin_razon_social(x.get("name"))),
            estado,
            titulo(x.get("city")),
            direccion,
            round(lat, 6),
            round(lng, 6),
            f"https://www.google.com/maps?q={lat},{lng}",
        ])

    print(f"  {len(filas)} con coordenadas")
    if sin_coords:
        print(f"  {sin_coords} sin coordenadas (Urrea los tiene en 0,0)")
    if fuera_mx:
        print(f"  {fuera_mx} con coordenada fuera de Mexico (descartados)")

    filas = resolver_genericos(filas)
    filas = sin_repetidos(filas)
    filas.sort(key=lambda f: (sin_acentos(f[0]).upper(), sin_acentos(f[4]).upper()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sucursales Urrea"
    ws.append(COLUMNAS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for f in filas:
        ws.append(f)
    for i, ancho in enumerate([46, 10, 34, 20, 26, 62, 12, 12, 52], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DESTINO)
    MARCA_GENERADO.write_text(str(int(DESTINO.stat().st_mtime)))

    print(f"\nListo: {DESTINO}")
    print(f"  filas: {len(filas)}   distribuidores: {len({f[0] for f in filas})}")


if __name__ == "__main__":
    main()
