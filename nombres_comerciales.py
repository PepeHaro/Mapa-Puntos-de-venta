#!/usr/bin/env python3
"""Cambia razones sociales por el nombre comercial en los siete archivos.

Uso:
    python3 nombres_comerciales.py            # muestra que cambiaria
    python3 nombres_comerciales.py --aplicar  # reescribe los Excel

Las marcas capturan al distribuidor como pueden: unas ponen el nombre con el
que rotula la tienda y otras la razon social o el apellido del dueno. El mismo
negocio termina partido en varias empresas.

`normalizar_nombres.py` resuelve lo mecanico (acentos, siglas legales, espacios).
Aqui van los que solo se resuelven sabiendo de quien se trata, y cada uno lleva
anotado con que se comprobo.

Tres formas de descubrirlos:

* **La sigla.** CERAMAT es CERAmica y MATeriales; MAPCO, MAPco COmerciales.
  Se confirma viendo cuantas sucursales caen encima de las que ya estaban.
* **El rotulo de la tienda.** Porcelanite guarda el nombre comercial en el
  titulo de cada sucursal aunque la categoria diga el apellido del dueno.
* **Buscarlo.** Para las razones sociales que no dicen nada, no queda mas que
  buscar la empresa y ver con que nombre opera.
"""

import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

SERVIDOR = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales"
)
FUENTES = [
    ("Distribuidores.xlsx", "BD"),
    ("Marcas/Daltile.xlsx", None),
    ("Marcas/Interceramic.xlsx", None),
    ("Marcas/Cesantoni.xlsx", None),
    ("Marcas/Vitromex.xlsx", None),
    ("Marcas/Porcelanite.xlsx", None),
    ("Marcas/Urrea.xlsx", None),
]

# razon social  ->  nombre comercial
COMERCIAL = {
    # La sigla, confirmada porque sus sucursales caen sobre las que ya estaban
    "CERAMICA Y MATERIALES CONTINENTAL": "CERAMAT",        # 87 sucursales a <3 km
    "MAPCO MATERIALES": "MAPCO",                           # 11
    "MATERAMA DE MEXICO": "MATERAMA",                      # 11
    "KURODA PLOMERIA Y AZULEJOS": "KURODA",
    "KURODA SAN": "KURODA",
    "MARIA ESTELA KURODA SAN": "KURODA",
    "BODEGA DE AZULEJOS Y PLOMERIA DALVI": "DALVI",        # 6

    # Mismo negocio, comprobado porque la mayoria de las tiendas del nombre
    # largo estan en la misma direccion que las del corto
    "DISTRIBUIDOR AZULEJERO DE MEXICO": "DAM",             # 25 direcciones iguales
    "NOVA CASA ZAMORA": "NOVACASA",                        # 10 de 13
    "LA INDUSTRIAL MEXICANA": "LIMSA",                     # 8 de 10
    "PLOMERIA GARCIA DE MONTERREY": "GRUPO PLOMERÍA GARCÍA",  # 10 de 14
    "MADERERIA ALIANZA": "CONSTRU ALIANZA",                # 8 de 13
    "KURODA NORTE": "KURODA",                              # 7 de 8
    "EXPO CERAMICAS JALISCO MX": "EXPOCERAMICAS",          # 7 de 9
    "EXPO CERAMICAS JALISCO": "EXPOCERAMICAS",
    "LA PALOMA PERIFÉRICO": "PISOS LA PALOMA",             # 4 de 5
    "FERRETERIA BALDOR": "BALDOR",
    "DISTRIBUIDORA ZIGA": "ZIMAT",
    "SANITARIOS AZULEJOS Y RECUBRIMIENTO": "SAR",
    "SANITARIOS AZULEJOS Y RECUBRIMIENTOS": "SAR",
    "CONTROL TECNICO DE FLUIDOS ARFI": "GRUPO ARFI",
    "AZULEJOS Y COMPLEMENTOS": "AZYCO",
    "EXCLUSIVA CERAMICA": "CERÁMICAS EXCLUSIVAS",
    "CONSTRUCTORA Y PROVEEDORA FERRETODO": "FERRETERO",
    "MATERIALES BUCIO Y YAÑEZ": "AZUMICH",
    "FERRETERIA Y MATERIALES LIZARRAGA": "GRUPO LIZARRAGA",
    "COMERCIALIZADORA SDMHC SA DE CV": "SODIMAC",
    "FERNANDO REQUEJO": "REQUEJO",
    "MARIO AARON SOTO CASTRO": "SOTO",
    "FERRETERIA AMAYA": "AMAYA",
    "NITROPISO AP": "TECNOPISO",
    "OUTLET KURODA": "KURODA",
    "OUTLET KURODA TRES RÍOS": "KURODA",
    "BODEGA KURODA": "KURODA",
    "LA PALOMA": "PISOS LA PALOMA",
    "RIVIERA AZULEJO Y MUEBLES": "AZULEJOS Y MUEBLES RIVIERA",
    "LA COMPETIDORA FERRETERAS.A.DEC.": "LA COMPETIDORA FERRETERA",
    "QUINCE ME S DE RL DE CV": "TERRATILE",
    "GILSA GARZA SADA": "GRUPO GILSA",
    # "Azulejera" es la palabra; la BD la traia con errata como "Azulejara"
    "AZULEJARA SAN JOSE": "AZULEJERA SAN JOSÉ",
    "AZULEJERA SAN JOSÉ MATRIZ": "AZULEJERA SAN JOSÉ",
    "AZULEJERA SAN JOSÉ BODEGA": "AZULEJERA SAN JOSÉ",
    "NITROPISO": "TECNOPISO",

    # Nombre comercial conocido por el negocio, no deducible del nombre legal
    "COMERCIALIZADORA INTEGRAL DE BAÑOS": "LLANO DE LA TORRE",
    "COMERCIALIZADORA INTEGRAL DE BAÑOS Y REVESTIMIENTOS": "LLANO DE LA TORRE",
    "COMPAÑIA MADERERA DE CHIHUAHUA SUCE": "HÁGALO",       # Hágalo Home Centers
    "COMPAÑIA MADERERA DE CHIHUAHUA SUCESORES": "HÁGALO",
}


# Cadenas que algunas fuentes capturan pegandole el nombre de la sucursal al de
# la empresa: "EL SURTIDOR - XOCHIMILCO", "VAMA JARIPILLO", "EL NIPLITO DEL
# SURESTE". El mismo negocio acaba partido en decenas de empresas de una tienda
# cada una. Todo lo que empiece asi es la misma cadena.
POR_PREFIJO = {
    "EL SURTIDOR -": "EL SURTIDOR",
    "GILSA": "GRUPO GILSA",
    "EL NIPLITO": "EL NIPLITO",
    "VAMA": "GRUPO VAMA",
    "GRUPO VAMA": "GRUPO VAMA",
}

# Empiezan igual pero son otro negocio.
#
# El Surtidor tiene nueve sucursales segun surtidor.com, en el centro del pais y
# Jalisco. "EL SURTIDOR DE PLOMERIA Y ..." son 32 tiendas de Yucatan y Quintana
# Roo: otra casa, que solo comparte las dos primeras palabras.
OTRA_EMPRESA = {
    "EL SURTIDOR DEL CONSTRUCTOR",
    "EL SURTIDOR PARA VIVIENDA",
    "EL SURTIDOR QUERETANO DEL SIGLO XXI",
    "EL SURTIDOR DE PLOMERIA Y",
    "EL SURTIDOR DE PLOMERIA Y MATERIALES",
}


def sin_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                   if unicodedata.category(c) != "Mn")


def clave(s):
    return " ".join(sin_acentos(str(s or "")).upper().split())


MAPA = {clave(k): v for k, v in COMERCIAL.items()}


def canonico(nombre):
    """Nombre comercial de una fila, o None si se queda como esta."""
    k = clave(nombre)
    if not k:
        return None
    if k in MAPA:
        return MAPA[k]
    if k in {clave(x) for x in OTRA_EMPRESA}:
        return None
    for prefijo, destino in POR_PREFIJO.items():
        p = clave(prefijo)
        if k == p or k.startswith(p + " ") or k.startswith(p + " -"):
            return destino
    return None


def main():
    aplicar = "--aplicar" in sys.argv

    libros, tocadas = {}, Counter()
    for archivo, hoja in FUENTES:
        ruta = SERVIDOR / archivo
        if not ruta.exists():
            continue
        libro = openpyxl.load_workbook(ruta)
        ws = libro[hoja] if hoja else libro[libro.sheetnames[0]]
        libros[archivo] = (libro, ruta, ws)
        for fila in ws.iter_rows(min_row=2):
            nuevo = canonico(fila[0].value)
            if nuevo and nuevo != str(fila[0].value):
                tocadas[(str(fila[0].value), nuevo)] += 1

    print(f"  cambios a hacer: {sum(tocadas.values())} filas\n")
    for (viejo, nuevo), n in tocadas.most_common():
        print(f"   {n:>4}  {viejo[:44]:46} -> {nuevo}")

    faltan = [k for k in COMERCIAL if clave(k) not in {clave(v) for v, _ in tocadas}]
    if faltan and not tocadas:
        print("\n  (ninguno aparece en los archivos)")

    if not aplicar:
        print("\n  Nada se escribio. Para aplicarlo:  python3 nombres_comerciales.py --aplicar")
        return

    total = 0
    for archivo, (libro, ruta, ws) in libros.items():
        cambios = 0
        for fila in ws.iter_rows(min_row=2):
            nuevo = canonico(fila[0].value)
            if nuevo and nuevo != str(fila[0].value):
                fila[0].value = nuevo
                cambios += 1
        if cambios:
            libro.save(ruta)
            total += cambios
    print(f"\n  {total} celdas actualizadas")


if __name__ == "__main__":
    main()
