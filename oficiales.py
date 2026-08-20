#!/usr/bin/env python3
"""Ajusta una cadena de la BD a las sucursales que declara su sitio oficial.

Uso:
    python3 oficiales.py            # muestra que quitaria
    python3 oficiales.py --aplicar  # borra las filas de mas

Cuando una marca reporta a un distribuidor, reporta las tiendas donde vende su
producto, y eso incluye locales que ya cerraron o que nunca fueron de la cadena.
Asi la BD acabo con 17 sucursales de DAM cuando su sitio lista 10, y 20 de
SANTANDREU cuando tiene 18.

Cada fila de la BD se empareja con la sucursal oficial mas parecida por
direccion. De cada sucursal oficial se conserva una sola fila, la mas completa;
las demas y las que no corresponden a ninguna se quitan.
"""

import math
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl

from limpiar_bd import calle, misma_tienda, na, numero_exterior, riqueza

BD = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales/Distribuidores.xlsx"
)

# (direccion, ciudad, estado) tal como las publica cada cadena.
OFICIALES = {
    # dampisos.com/sucursales
    "DAM": [
        ("Carretera Tepic-Gdl 606, Camichín de Jauja, 63508", "Tepic", "Nayarit"),
        ("Blvd Luis Donaldo Colosio 1496, Jacarandas, 63175", "Tepic", "Nayarit"),
        ("Calz. Jesús González Gallo 1174, Atlas, 44898", "Guadalajara", "Jalisco"),
        ("Calz. Jesús González Gallo 2501, El Rosario, 44898", "Guadalajara", "Jalisco"),
        ("Calz. Jesús González Gallo 290, San Carlos, 44460", "Guadalajara", "Jalisco"),
        ("Blvd Riviera Nayarit 14, 63735 Tondoroque", "Bahía de Banderas", "Nayarit"),
        ("Blvd. Tepic Xalisco 436, Los Fresnos, 63197", "Tepic", "Nayarit"),
        ("Av Niños Héroes 1055, Moderna, 44190", "Guadalajara", "Jalisco"),
        ("Av Niños Héroes 2365, Moderna, 44190", "Guadalajara", "Jalisco"),
        ("Juan Gil Preciado 4621, Nuevo México, 45140", "Zapopan", "Jalisco"),
    ],
    # santandreu.com.mx/sucursales
    "SANTANDREU": [
        ("Marcelino García Barragán, Tamulté de las Barrancas, 86150", "Villahermosa", "Tabasco"),
        ("Av. Aluminio 101-B, Ciudad Industrial, 86010", "Villahermosa", "Tabasco"),
        ("Av. Gregorio Méndez Magaña 830, Centro, 86000", "Villahermosa", "Tabasco"),
        ("Av. Constitución 1303, Centro, 86000", "Villahermosa", "Tabasco"),
        ("Prol. Paseo Usumacinta 4663, Guayabal, 86090", "Villahermosa", "Tabasco"),
        ("Boulevard Adolfo López Mateos s/n, Centro, 86323", "Comalcalco", "Tabasco"),
        ("Carr. Circuito del Golfo 407, Pueblo Nuevo, 86500", "Cárdenas", "Tabasco"),
        ("Cda. Abraham Bandala 305, Centro, 86560", "Cárdenas", "Tabasco"),
        ("Av Periférico S/N, Centro, 86690", "Cunduacán", "Tabasco"),
        ("Calle Venustiano Carranza 85, Enrique Gonzalez Pedrero, 86200", "Jalpa de Méndez", "Tabasco"),
        ("Carlos A. Madrazo 60, Independencia, 86720", "Macuspana", "Tabasco"),
        ("Benito Juárez 608, Centro, 86600", "Paraíso", "Tabasco"),
        ("Paraíso - Comalcalco, 86608", "Paraíso", "Tabasco"),
        ("Galeana S/N, Centro, 86500", "Cárdenas", "Tabasco"),
        ("Av Justo Sierra 156, Fraccionamiento Nuevo, 96734", "Minatitlán", "Veracruz"),
        ("Av Universidad Veracruzana Km 7+600, Santa Cecilia, 96510", "Coatzacoalcos", "Veracruz"),
        ("Av Patricio Trueba de Regil SN, Sascalum, 24095", "San Francisco de Campeche", "Campeche"),
        ("Puerto Real S/N, Luis Donaldo Colosio Murrieta, 24110", "Ciudad del Carmen", "Campeche"),
    ],
}


def km(a, b, c, d):
    return (((a - c) * 110.57) ** 2
            + ((b - d) * 111.32 * math.cos(math.radians(a))) ** 2) ** 0.5


def parecido(fila, oficial):
    """Que tanto se parece una fila de la BD a una sucursal oficial."""
    señas = (str(fila[6].value or ""), str(fila[5].value or ""), str(fila[4].value or ""))
    if misma_tienda(señas, oficial, 0):
        return 3
    na_, nb = numero_exterior(*señas), numero_exterior(*oficial)
    comun = calle(*señas) & calle(*oficial)
    if na_ is not None and na_ == nb and comun:
        return 3
    if comun and na(señas[1]) == na(oficial[1]):
        return 2
    if comun:
        return 1
    return 0


def main():
    aplicar = "--aplicar" in sys.argv
    libro = openpyxl.load_workbook(BD)
    hoja = libro["BD"]

    fuera = []
    for cadena, sucursales in OFICIALES.items():
        filas = [f for f in hoja.iter_rows(min_row=2)
                 if f[0].value and " ".join(str(f[0].value).split()) == cadena]
        print(f"\n=== {cadena}: la BD tiene {len(filas)}, el sitio oficial lista {len(sucursales)} ===")

        # Emparejar una por una no alcanza: la misma tienda aparece con calles
        # distintas segun quien la capturo. La "Periférico 2020" de Villahermosa
        # es la que el sitio llama "Marcelino García Barragán", mismo codigo
        # postal. Asi que se ordenan por parecido y se conservan tantas como
        # sucursales declare la cadena.
        puntuadas = []
        for f in filas:
            mejor = max((parecido(f, s), i) for i, s in enumerate(sucursales))
            puntuadas.append((mejor[0], riqueza(f), f, sucursales[mejor[1]]))
        puntuadas.sort(key=lambda x: (-x[0], -x[1]))

        for punt, _, f, cual in puntuadas[len(sucursales):]:
            fuera.append(f[0].row)
            print(f"   quita  fila {f[0].row:>5}  {str(f[5].value)[:14]:16} {str(f[6].value)[:48]}")
        quedan = puntuadas[:len(sucursales)]
        print(f"   se quedan {len(quedan)}; la peor coincidencia puntua {quedan[-1][0] if quedan else 0}")

    print(f"\n  filas a quitar: {len(fuera)}")
    if not aplicar:
        print("  Nada se escribio. Para aplicarlo:  python3 oficiales.py --aplicar")
        return

    for r in sorted(set(fuera), reverse=True):
        hoja.delete_rows(r)
    for n in range(2, hoja.max_row + 1):
        hoja.cell(n, 10, f'=IF(H{n}="","","https://www.google.com/maps?q="&H{n}&","&I{n})')
    for i in range(6):
        try:
            libro.save(BD)
            break
        except OSError:
            if i == 5:
                raise
            time.sleep(4)
    print(f"  BD queda en {hoja.max_row - 1} filas")


if __name__ == "__main__":
    main()
