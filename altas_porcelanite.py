#!/usr/bin/env python3
"""Da de alta en la BD los distribuidores que solo trae Porcelanite.

Uso:
    python3 altas_porcelanite.py            # muestra que agregaria
    python3 altas_porcelanite.py --aplicar  # escribe en Distribuidores.xlsx

Regla, la misma que se uso con Vitromex: si el distribuidor ya existe en la BD
como grupo, todas sus sucursales ya estan y no se toca nada. Solo entran
empresas que la BD no tiene.

Antes de dar de alta, cada sucursal se compara contra la BD. La cercania sola
no sirve como criterio: la 25 Poniente de Puebla es un corredor de azulejeras
pegadas, y tres negocios distintos caben en 40 metros. Lo que decide es el
numero exterior de la calle, igual que en el resto del proyecto.
"""

import math
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

SERVIDOR = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales"
)
BD = SERVIDOR / "Distribuidores.xlsx"
PORCELANITE = SERVIDOR / "Marcas" / "Porcelanite.xlsx"

# Comparten direccion exacta con una empresa de la BD, pero no alcanzo para
# confirmar que sean la misma razon social. Se quedan fuera: es preferible que
# falte un distribuidor a que se duplique uno.
DUDOSOS = {"GRUPO JCC", "VIZUET", "GUTIERREZ", "PALENCIA", "BERNARDO MENESES"}

RUIDO = {"grupo", "comercializadora", "distribuidora", "distribuidor",
         "corporativo", "sa", "de", "cv", "s", "rl", "sapi", "gpo"}

CERCA_M = 150  # solo se comparan direcciones dentro de este radio

# Palabras que no distinguen una direccion de otra.
VIAS = {"av", "avenida", "blvd", "boulevard", "calle", "calz", "calzada", "carr",
        "carretera", "prol", "prolongacion", "col", "colonia", "int", "num",
        "numero", "esq", "local", "lote", "sur", "norte", "ote", "pte",
        "centro", "fracc", "esquina"}


def sin_cp(direccion):
    """Quita el codigo postal, que si no se cuela como numero de la calle.

    Hay que buscarlo por la etiqueta "CP" y no solo por largo: los codigos de
    la Ciudad de Mexico empiezan con cero y el Excel se lo come, asi que
    03300 llega escrito como 3300 y pasaria por numero exterior.
    """
    t = na(direccion)
    t = re.sub(r"\bc\s*p\s*:?\s*\d{4,5}\b", " ", t)
    t = re.sub(r"\b\d{5}\b", " ", t)
    return t


def numero_exterior(direccion):
    """El numero mas grande de la direccion, que casi siempre es el del local.

    En Puebla las calles se llaman "25 Poniente", asi que la direccion trae dos
    numeros y el del local es el mayor: en "25 Poniente 3504" el local es 3504.
    """
    nums = [int(n) for n in re.findall(r"\b(\d{1,5})\b", sin_cp(direccion))]
    return max(nums) if nums else None


def calle(direccion):
    return {w for w in re.sub(r"\d+", " ", sin_cp(direccion)).split()
            if len(w) > 3 and w not in VIAS}


def misma_tienda(dir_a, dir_b):
    """Misma direccion: mismo numero exterior y alguna palabra de calle en comun."""
    na_, nb = numero_exterior(dir_a), numero_exterior(dir_b)
    if na_ is None or nb is None or na_ != nb:
        return False
    return bool(calle(dir_a) & calle(dir_b))


def na(s):
    s = "".join(c for c in unicodedata.normalize("NFD", str(s or "").lower())
                if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def clave(s):
    """Nombre reducido a sus palabras distintivas, para comparar razones sociales."""
    return " ".join(sorted(w for w in re.sub(r"[^\w\s]", " ", na(s)).split()
                           if w not in RUIDO and len(w) > 2))


def km(a, b, c, d):
    return (((a - c) * 110.57) ** 2
            + ((b - d) * 111.32 * math.cos(math.radians(a))) ** 2) ** 0.5


def main():
    aplicar = "--aplicar" in sys.argv

    libro = openpyxl.load_workbook(BD)  # sin data_only: conserva las formulas
    hoja = libro["BD"]
    existentes, puntos = set(), defaultdict(list)
    for fila in hoja.iter_rows(min_row=2):
        if not fila[0].value:
            continue
        existentes.add(clave(fila[0].value))
        try:
            lat, lon = float(fila[7].value), float(fila[8].value)
        except (TypeError, ValueError):
            continue
        puntos[(round(lat, 2), round(lon, 2))].append((lat, lon, str(fila[6].value or "")))

    wb = openpyxl.load_workbook(PORCELANITE, read_only=True, data_only=True)
    porcelanite = [list(f) for f in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True)]
    wb.close()

    candidatas = [f for f in porcelanite
                  if clave(f[0]) not in existentes and f[0] not in DUDOSOS]

    altas, repetidas = [], 0
    for f in candidatas:
        lat, lon = f[6], f[7]
        vecinas = [
            d for dla in (-0.01, 0, 0.01) for dlo in (-0.01, 0, 0.01)
            for x, y, d in puntos.get((round(lat + dla, 2), round(lon + dlo, 2)), [])
            if km(lat, lon, x, y) * 1000 < CERCA_M
        ]
        if any(misma_tienda(f[5], d) for d in vecinas):
            repetidas += 1
        else:
            altas.append(f)

    ya = len({f[0] for f in porcelanite}) - len({f[0] for f in candidatas}) - len(
        {f[0] for f in porcelanite if f[0] in DUDOSOS})
    print(f"  distribuidores que la BD ya tiene:  {ya:>3}  (no se tocan)")
    print(f"  dudosos, se dejan fuera:            {len(DUDOSOS):>3}")
    print(f"  ya estaban en la BD con otro nombre: {repetidas:>3}  (misma direccion, se descartan)")
    print(f"  ALTAS: {len(altas)} sucursales de {len({f[0] for f in altas})} distribuidores")

    if not aplicar:
        print("\n  Nada se escribio. Para aplicarlo:  python3 altas_porcelanite.py --aplicar")
        return

    n = hoja.max_row
    for f in altas:
        n += 1
        hoja.cell(n, 1, f[0])   # DISTRIBUIDOR
        hoja.cell(n, 2, f[2])   # NOMBRE
        hoja.cell(n, 4, f[1])   # PAÍS
        hoja.cell(n, 5, f[3])   # ESTADO
        hoja.cell(n, 6, f[4])   # CIUDAD
        hoja.cell(n, 7, f[5])   # DIRECCIÓN
        hoja.cell(n, 8, f[6])   # LATITUD
        hoja.cell(n, 9, f[7])   # LONGITUD
        hoja.cell(n, 10, f'=IF(H{n}="","","https://www.google.com/maps?q="&H{n}&","&I{n})')
    libro.save(BD)
    print(f"\n  BD queda en {hoja.max_row - 1} filas")


if __name__ == "__main__":
    main()
