#!/usr/bin/env python3
"""Deja un solo nombre por distribuidor en los siete archivos.

Uso:
    python3 normalizar_nombres.py            # muestra que cambiaria
    python3 normalizar_nombres.py --aplicar  # reescribe los Excel

El mismo negocio esta escrito de varias formas segun quien lo capturo, y cada
forma cuenta como una empresa distinta: se parten sus sucursales en el mapa y
sus tiendas nunca se comparan entre si para detectar duplicados. "PISOS IBERIA"
y "PISOS IBERIA " sumaban 23 sucursales cuando la cadena tiene 10.

Solo se unen los casos donde la diferencia no deja lugar a dudas:

* espacios de mas, acentos, puntuacion o mayusculas
* las siglas de la forma legal al final (SA de CV, S de RL de CV)
* el "Grupo", "Corporativo" o "Comercializadora" del principio

No se tocan los que se parecen pero podrian ser otra empresa. "CERÁMICA" y
"CERÁMICA SJ" siguen separadas, igual que "FERRETERO" y "GRUPO FERRETERO RC".
Esos hay que resolverlos viendo la direccion, no el nombre.
"""

import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SERVIDOR = Path(
    "/Volumes/MK/MK Server/Finanzas/Strategy 2026/Distribuidores/Data/Sucursales"
)
# (archivo, hoja, columna del distribuidor)
FUENTES = [
    ("Distribuidores.xlsx", "BD", 1),
    ("Marcas/Daltile.xlsx", None, 1),
    ("Marcas/Interceramic.xlsx", None, 1),
    ("Marcas/Cesantoni.xlsx", None, 1),
    ("Marcas/Vitromex.xlsx", None, 1),
    ("Marcas/Porcelanite.xlsx", None, 1),
    ("Marcas/Urrea.xlsx", None, 1),
]

# Palabras de figura societaria que van al principio y no distinguen al negocio.
PREFIJOS = ("grupo", "gpo", "corporativo", "comercializadora", "distribuidora",
            "comercial", "importadora")


def sin_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                   if unicodedata.category(c) != "Mn")


def base(nombre):
    """Nombre reducido a lo que de verdad lo distingue."""
    t = sin_acentos(nombre).lower()
    t = " ".join(re.sub(r"[^\w\s]", " ", t).split())
    t = re.sub(r"\s+s\s*a\s*p\s*i(\s+de)?(\s+c\s*v)?\s*$", "", t)
    t = re.sub(r"\s+s(\s+de)?\s+r\s*l(\s+de)?(\s+c\s*v)?\s*$", "", t)
    t = re.sub(r"\s+s\s*a(\s+de?)?(\s+c\s*v?)?\s*$", "", t)
    t = re.sub(r"\s+de\s+c\s*v\s*$", "", t)
    return t.strip()


def raiz(nombre):
    """Como `base`, pero ademas sin el prefijo societario."""
    t = base(nombre)
    partes = t.split()
    while partes and partes[0] in PREFIJOS:
        partes = partes[1:]
    return " ".join(partes) or t


# Caracteres que solo aparecen cuando el acento se guardo mal: "FERREBAŃOS"
# en vez de "FERREBAÑOS".
ROTOS = set("ŃńÃÂ¿½Ā")


def puntaje(nombre, variantes, en_bd):
    """Cual de las formas se queda. Gana, en este orden:

    1. la que no trae el acento mal guardado
    2. la que no arrastra las siglas de la forma legal
    3. la que mas usa la BD, que es el archivo de referencia
    4. la mas repetida en el resto de los archivos
    """
    return (
        not (set(nombre) & ROTOS),
        base(nombre) == " ".join(re.sub(r"[^\w\s]", " ", sin_acentos(nombre).lower()).split()),
        en_bd.get(nombre, 0),
        variantes[nombre],
        -len(nombre),
    )


def main():
    aplicar = "--aplicar" in sys.argv

    libros, usos = {}, defaultdict(Counter)
    en_bd = Counter()
    for archivo, hoja, col in FUENTES:
        ruta = SERVIDOR / archivo
        if not ruta.exists():
            continue
        libro = openpyxl.load_workbook(ruta)
        ws = libro[hoja] if hoja else libro[libro.sheetnames[0]]
        libros[archivo] = (libro, ruta, ws, col)
        for fila in ws.iter_rows(min_row=2):
            v = fila[col - 1].value
            if not v:
                continue
            nombre = " ".join(str(v).split())
            usos[raiz(nombre)][nombre] += 1
            if hoja == "BD":
                en_bd[nombre] += 1

    grupos = {r: v for r, v in usos.items() if r and len(v) > 1}

    canon = {}
    for r, variantes in grupos.items():
        elegido = max(variantes, key=lambda n: puntaje(n, variantes, en_bd))
        for n in variantes:
            if n != elegido:
                canon[n] = elegido

    filas = sum(c for r in grupos for c in usos[r].values())
    print(f"  grupos con mas de una forma de escribirse: {len(grupos)}")
    print(f"  nombres que cambian: {len(canon)}   filas afectadas: {filas}\n")
    for r, variantes in sorted(grupos.items(), key=lambda x: -sum(x[1].values())):
        elegido = max(variantes, key=lambda n: puntaje(n, variantes, en_bd))
        otras = [f"{n!r}({variantes[n]})" for n in variantes if n != elegido]
        print(f"   {sum(variantes.values()):>4}  {elegido!r}  <-  " + ", ".join(otras)[:88])

    if not aplicar:
        print("\n  Nada se escribio. Para aplicarlo:  python3 normalizar_nombres.py --aplicar")
        return

    cambios = 0
    for archivo, (libro, ruta, ws, col) in libros.items():
        tocado = False
        for fila in ws.iter_rows(min_row=2):
            v = fila[col - 1].value
            if not v:
                continue
            nombre = " ".join(str(v).split())
            nuevo = canon.get(nombre, nombre)
            if nuevo != str(v):
                fila[col - 1].value = nuevo
                cambios += 1
                tocado = True
        if tocado:
            libro.save(ruta)
    print(f"\n  {cambios} celdas actualizadas en {len(libros)} archivos")


if __name__ == "__main__":
    main()
